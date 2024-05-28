import openpyxl as opx
import tkinter as tk
from tkinter import ttk
from tkinter import *
import datetime
from pandas import ExcelWriter
import pandas as pd
from ttkthemes import ThemedTk
from tkinter import messagebox



filepath = r"C:\Users\borko.kovacevic\Desktop\Project\evPR.xlsx" 
today = datetime.datetime.today().strftime('%d.%m.%Y') 
z = 0
z += 1

file = r"C:\Users\borko.kovacevic\Desktop\Project\evPR.xlsx"
wb = opx.load_workbook(file)
ws = wb["Ljudi"]
names = [ws.cell(row=i,column=1).value for i in range(2,ws.max_row+1)]
# Create a list to hold the comboboxes
combos = []
wsF = wb["FIRME"]
companies = [wsF.cell(row=i,column=1).value for i in range(2,wsF.max_row+1)]
comp = []

def filter_names(event, combo):
    # Get the typed text from the combobox
    typed_text = combo.get()

    # Filter the names list based on the typed text
    filtered_names = [name for name in names if typed_text.lower() in name.lower()]

    # Update the combobox dropdown list with the filtered names
    combo['values'] = filtered_names


class Nalog:
    def __init__(self, br_naloga, sektor, izdavalac, primalac, clan1, clan2, clan3, clan4, klijent, 
                 mesto, adresa, kontakt, aktivnost, datum_izd, 
                 datum_rok, napomena, br_ponude, br_doc) -> None:
        self.br_naloga = br_naloga
        self.sektor = sektor
        self.izdavalac = izdavalac
        self.primalac = primalac
        self.clan1 = clan1
        self.clan2 = clan2
        self.clan3 = clan3
        self.clan4 = clan4
        self.klijent = klijent
        self.mesto  = mesto
        self.adresa = adresa
        self.kontakt = kontakt
        self.aktivnost = aktivnost
        self.datum_izd = datum_izd
        self.datum_rok = datum_rok
        self.napomena = napomena
        self.br_ponude = br_ponude
        self.br_doc = br_doc
    
    @classmethod
    def unesi(self):
        global z
        z += 1
        zaunos = (int(no_nalog_entry.get()), sektor_cb.get(), nalog_izdao_entry.get(), nalog_primo_vodja_entry.get(), 
                  nalog_primo_clan1_entry.get(), nalog_primo_clan2_entry.get(), nalog_primo_clan3_entry.get(),nalog_primo_clan4_entry.get(), 
                  klijent_entry.get(), mesto_entry.get(), adresa_entry.get(), kontakt_Enry.get(),
                    zadatak_entry.get(), D_izdavanja_Entry.get(), rok_Entry.get(), napomena_entry.get(), 
                    Ponuda_Entry.get(), isprava_Entry.get())
        
        workbook = opx.load_workbook(filepath)
        sheet = workbook.active
        sheet.append(zaunos)
        
        isprava_Entry.delete(0, END)
        isprava_Entry.insert(0, no_nalog_entry.get() + "." + sektor_cb.get() + "." + str(z))
    
        workbook.save(filepath)
    @classmethod
    def upamti_klijenta(self):
        

        zaunos = (klijent_entry.get(), Ponuda_Entry.get(), kontakt_Enry.get(), mesto_entry.get(), adresa_entry.get())
        
        workbook = opx.load_workbook(filepath)
        sheet = workbook["FIRME"]
        sheet.append(zaunos)
       
        workbook.save(filepath)


def sektor_changed(event):
   
    
    AB = "Anđelko Baskić"
    ML = "Milomir Lukić"
    MT = "Milan Trišić"
    MB = "Miodrag Brklje"
    MV = "Milica Vučićević"
    SVI = "Nemanja Mitrović, Lazar Lešnjak, Dragana Mudavdžić"
    isprava_Entry.delete(0, END)
    isprava_Entry.insert(0, no_nalog_entry.get() + "." + sektor_cb.get() + "." + str(z))

    if sektor_cb.get() == "KT":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, AB + ", " + ML)
    elif sektor_cb.get() == "SPI":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, MB + ", " + MT)
    elif sektor_cb.get() == "LB":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, MB + ", " + MT)
    elif sektor_cb.get() == "CE":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, MV)
    elif sektor_cb.get() == "SSP":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, SVI)       
    elif sektor_cb.get() == "SV":
        nalog_izdao_entry.delete(0, END)
        nalog_izdao_entry.insert(0, MT)


def close_app():
    if messagebox.askokcancel("ZATVARANJE", "Da li želite da zatvorite aplikaciju?"):
        window.destroy()

def broj_ponude(event):
    
    firme_df = pd.read_excel(filepath, sheet_name="FIRME")

    subklijent = firme_df.loc[firme_df['FIRMA'] == klijent_entry.get()]
    if len(subklijent) == 0:
        Ponuda_Entry.delete(0, END)
        Ponuda_Entry.insert(0, "br. ponude od (datum)")
        kontakt_Enry.delete(0, END)
        kontakt_Enry.insert(0, "ime i prezime, br. telefona")
        mesto_entry.delete(0, END)
        mesto_entry.insert(0, "/")
        adresa_entry.delete(0, END)
        adresa_entry.insert(0, "/")    
    else:
        Ponuda_Entry.delete(0, END)
        Ponuda_Entry.insert(0, subklijent.values[0, 1])
        kontakt_Enry.delete(0, END)
        kontakt_Enry.insert(0, subklijent.values[0, 2])
        mesto_entry.delete(0, END)
        mesto_entry.insert(0, subklijent.values[0, 3])
        adresa_entry.delete(0, END)
        adresa_entry.insert(0, subklijent.values[0, 4])          
        klijent_entry.delete(0, END)
        klijent_entry.insert(0, subklijent.values[0, 5])  

def reset():
    global z
    z=1
    workbook = opx.load_workbook(filepath)
    sheet = workbook.active
    sektor_cb.delete(0, END)
    no_nalog_entry.delete(0, END)
    no_nalog_entry.insert(0, sheet[f"A{sheet.max_row}"].value + 1)
    nalog_izdao_entry.delete(0, END)
    nalog_primo_vodja_entry.delete(0, END)
    nalog_primo_clan1_entry.delete(0, END)
    nalog_primo_clan2_entry.delete(0, END)
    nalog_primo_clan3_entry.delete(0, END)
    nalog_primo_clan4_entry.delete(0, END)
    kontakt_Enry.delete(0, END)
    klijent_entry.delete(0, END)
    mesto_entry.delete(0, END)
    adresa_entry.delete(0, END)
    napomena_entry.delete(0, END)
    zadatak_entry.delete(0, END)
    D_izdavanja_Entry.delete(0, END)
    D_izdavanja_Entry.insert(0, today)
    rok_Entry.delete(0, END)
    rok_Entry.insert(0, today)
    isprava_Entry.delete(0, END)
    Ponuda_Entry.delete(0, END)




def k1():
    k_1 = "Periodično Kontrolisanje instalacija hidrantske mreže za gašenje požara"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_1)

def k2():
    k_2 = "Periodično Kontrolisanje mobilnih uređaja za gašenje požara"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_2)

def k3():
    k_3 = "Periodično Kontrolisanje instalacija i uređaja za automatsko otkrivanje i dojavu požara"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_3)

def k4():
    k_4 = "Periodično Kontrolisanje instalacija i uređaja za gašenje požara"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_4)

def k5():
    k_5 = "Periodično Kontrolisanje instalacija i uređaja za detekciju eksplozivnih i zapaljivih gasova"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_5)

def k6():
    k_6 = "Periodično Kontrolisanje instalacija za odvođenje dima i toplote"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_6)

def k7():
    k_7 = "Periodično Kontrolisanje instalacija i uređaja u zonama opasnosti od eksplozije"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, k_7)

def im1():
    im_1 = "Ispitivanje mikroklime u zimskom periodu i osvetljenosti"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_1)

def im2():
    im_2 = "Ispitivanje mikroklime u letnjem periodu i osvetljenosti"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_2)

def im3():
    im_3 = "Pregled i ispitivanje gromobranskih instalacija"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_3)

def im4():
    im_4 = "Pregled i ispitivanje električnih instalacija"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_4)

def im5():
    im_5 = "Kontrolisanje PP rasvete"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_5)

def lb1():
    im_1 = "Ispitivanje izolacionih odela i aparata"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_1)

def lb2():
    im_2 = "Ispitivanje hidrantskih creva"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_2)

def lb3():
    im_3 = "prašina"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_3)

def lb4():
    im_4 = "još nešto akreditovano"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_4)

def lb5():
    im_5 = "i još nešto akreditovano"
    zadatak_entry.delete(0, END)
    zadatak_entry.insert(0, im_5)


#user forme
window = ThemedTk(theme="arc", background="yellow")
stil = ttk.Style(window)
stil.theme_use("arc")
window.title("NALOG ZA RAD")
#window.eval("tk::PlaceWindow . topleft")
window.iconbitmap(r"\\fileserver\TEHPRO\ZAŠTITA\TEHNIČKI POSLOVI\EVIDENCIJE\Evidencije RN\PROBA\assets\Tehpro-beli1.ico")
#window.geometry("1470x650")
ram = tk.Frame(window)
ram.pack()
  
basic_info_nalog =tk.LabelFrame(ram, text="OSNOVNI PODACI")
basic_info_nalog.grid(row= 0, column=0, sticky="news", padx=20, pady=10)

D_izdavanja_Label =tk.Label(basic_info_nalog, text="Datum izdavanja:" )
D_izdavanja_Label.grid(row=1, column=2, padx=5, pady=5, sticky="e")
D_izdavanja_Entry =ttk.Entry(basic_info_nalog, width=30)
D_izdavanja_Entry.grid(row=1, column=3)
D_izdavanja_Entry.insert(0, today)

Ponuda_Label =tk.Label(basic_info_nalog, text="Br. ponude:")
Ponuda_Label.grid(row=0, column=4, padx=5, pady=5, sticky="ew")
Ponuda_Entry =ttk.Entry(basic_info_nalog, width=30)
Ponuda_Entry.grid(row=1, column=4)

isprava_Label =tk.Label(basic_info_nalog, text="Br. dokumenta:", width=15 )
isprava_Label.grid(row=2, column=4, padx=5, pady=5, sticky="ew")
isprava_Entry =ttk.Entry(basic_info_nalog, width=15)
isprava_Entry.grid(row=3, column=4)

rok_label = tk.Label(basic_info_nalog, text="Rok za izvršenje:" )
rok_label.grid(row=2, column=2, padx=5, pady=5, sticky="e")
rok_Entry = ttk.Entry(basic_info_nalog, width=30)
rok_Entry.grid(row=2, column=3, padx=5, pady=5)
rok_Entry.insert(0, today)

kontakt_label = tk.Label(basic_info_nalog, text="Kontakt osoba" )
kontakt_label.grid(row=3, column=2, padx=5, pady=5, sticky="e")
kontakt_Enry = ttk.Entry(basic_info_nalog, width=30)
kontakt_Enry.grid(row=3, column=3, padx=5, pady=5)

sektor_label =tk.Label(basic_info_nalog, text="SEKTOR", width=10)
sektor_cb = ttk.Combobox(basic_info_nalog, values=["KT", "LB", "SPI", "SSP", "CE", "SV"], width=10)
sektor_label.grid(row=0, column=0, sticky="w")
sektor_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")

no_nalog_label=tk.Label(basic_info_nalog, text="Br naloga", width=10 )
no_nalog_label.grid(row=1, column=0, padx=2, pady=2, sticky="w")
nalog_izdao_label=tk.Label(basic_info_nalog, text="Nalog izdao", width=10)
nalog_izdao_label.grid(row=2, column=0, padx=2, pady=2, sticky="w")

no_nalog_entry = ttk.Entry(basic_info_nalog, width=10, font=["center", 10, "bold"])
wb = opx.load_workbook(filepath)
sheet = wb.active
if sheet[f"A{sheet.max_row}"].value == "Br.Naloga" or sheet[f"A{sheet.max_row}"].value == None:
       no_nalog_entry.insert(0, "1")
        
else:
           no_nalog_entry.insert(0, sheet[f"A{sheet.max_row}"].value + 1)

nalog_izdao_entry = ttk.Entry(basic_info_nalog, width=30)
no_nalog_entry.grid(row=1, column=1, sticky="w")
nalog_izdao_entry.grid(row=2, column=1, sticky="w")

nalog_primo_vodja_label = tk.Label(basic_info_nalog, text="Nalog primio", width=10)
nalog_primo_vodja_label.grid(row=3, column=0, padx=2, pady=2, sticky="w")

nalog_primo_vodja_entry = ttk.Combobox(basic_info_nalog, width=25)
nalog_primo_vodja_entry.grid(row=3, column=1, padx=2, pady=2)
nalog_primo_vodja_entry.bind("<KeyRelease>", lambda event: filter_names(event, nalog_primo_vodja_entry))
combos.append(nalog_primo_vodja_entry)

nalog_primo_clan1_label = tk.Label(basic_info_nalog, text="Članovi tima", width=10)
nalog_primo_clan1_label.grid(row=4, column=0, padx=2, pady=2, sticky="w")

nalog_primo_clan1_entry = ttk.Combobox(basic_info_nalog, width=25)
nalog_primo_clan1_entry.grid(row=4, column=1, padx=2, pady=2)
nalog_primo_clan1_entry.bind("<KeyRelease>", lambda event: filter_names(event, nalog_primo_clan1_entry))
combos.append(nalog_primo_clan1_entry)
nalog_primo_clan2_entry = ttk.Combobox(basic_info_nalog, width=25)
nalog_primo_clan2_entry.grid(row=4, column=2, padx=2, pady=2)
nalog_primo_clan2_entry.bind("<KeyRelease>", lambda event: filter_names(event, nalog_primo_clan2_entry))
combos.append(nalog_primo_clan2_entry)
nalog_primo_clan3_entry = ttk.Combobox(basic_info_nalog, width=25)
nalog_primo_clan3_entry.grid(row=4, column=3, padx=2, pady=2)
nalog_primo_clan3_entry.bind("<KeyRelease>", lambda event: filter_names(event, nalog_primo_clan3_entry))
combos.append(nalog_primo_clan3_entry)
nalog_primo_clan4_entry = ttk.Combobox(basic_info_nalog, width=25)
nalog_primo_clan4_entry.grid(row=4, column=4, padx=2, pady=2)
nalog_primo_clan4_entry.bind("<KeyRelease>", lambda event: filter_names(event, nalog_primo_clan4_entry))
combos.append(nalog_primo_clan4_entry)

# drugi deo naloga sa podacima o zadatku
zadatak_info_nalog =tk.LabelFrame(ram, text="PODACI O ZADATKU")
zadatak_info_nalog.grid(row= 1, column=0, sticky="news", padx=20, pady=10)

klijent_label = tk.Label(zadatak_info_nalog, text="Klijent")
klijent_label.grid(row=0, column=0, padx=2, pady=6, sticky="w")
klijent_entry = ttk.Entry(zadatak_info_nalog, width=30)
klijent_entry.grid(row=0, column=1, padx=5, pady=6)


mesto_label = tk.Label(zadatak_info_nalog, text="Mesto")
mesto_label.grid(row=1, column=0, padx=2, pady=6, sticky="w")
mesto_entry = ttk.Entry(zadatak_info_nalog, width=30)
mesto_entry.grid(row=1, column=1, padx=2, pady=6)

adresa_label = tk.Label(zadatak_info_nalog, text="Adresa")
adresa_label.grid(row=0, column=3, padx=5, pady=6, sticky="e")
adresa_entry = ttk.Entry(zadatak_info_nalog, width=30)
adresa_entry.grid(row=0, column=4, padx=2, pady=6)

napomena_label = tk.Label(zadatak_info_nalog, text="Napomena")
napomena_label.grid(row=1, column=3, padx=5, pady=6, sticky="e")
napomena_entry = ttk.Entry(zadatak_info_nalog, width=30)
napomena_entry.grid(row=1, column=4, padx=2, pady=6)

KT_FRAME =ttk.LabelFrame(ram, text="KONTROLNO TELO", width=400)
KT_FRAME.grid(row= 0, column=1, sticky="news", padx=20, pady=10)

sektor_cb.bind('<<ComboboxSelected>>', sektor_changed)
klijent_entry.bind("<FocusOut>", broj_ponude)

KT1 = ttk.Button(KT_FRAME, text="K - 1", command=k1, width=8)
KT1.grid(row=0, column=0, padx=10, pady=10)
KT2 = ttk.Button(KT_FRAME, text="K - 2", command=k2, width=8)
KT2.grid(row=0, column=1, padx=10, pady=10, )
KT3 = ttk.Button(KT_FRAME, text="K - 3",  command=k3, width=8)
KT3.grid(row=0, column=2, padx=10, pady=10, )
KT4 = ttk.Button(KT_FRAME, text="K - 4", command=k4, width=8)
KT4.grid(row=1, column=0, padx=10, pady=10, )
KT5 = ttk.Button(KT_FRAME, text="K - 5",  command=k5, width=8)
KT5.grid(row=1, column=1, padx=10, pady=10, )
KT6 = ttk.Button(KT_FRAME, text="K - 6", command=k6, width=8)
KT6.grid(row=1, column=2, padx=10, pady=10, )
KT7 = ttk.Button(KT_FRAME, text="K - 7",  command=k7, width=8)
KT7.grid(row=2, column=0, padx=10, pady=10)
#KT8 = ttk.Button(KT_FRAME, text="K - 8", width=8)
#KT8.grid(row=1, column=3, padx=10, pady=10)

IM_FRAME =ttk.LabelFrame(ram, text="SEKTOR PREGLEDA I ISPITIVANJA", width=400)
IM_FRAME.grid(row= 1, column=1, sticky="news", padx=20, pady=10)

IM1 = ttk.Button(IM_FRAME, text="SPI - 1", command=im1, width=8)
IM1.grid(row=0, column=0, padx=10, pady=10, )
IM2 = ttk.Button(IM_FRAME, text="SPI - 2",  command=im2, width=8)
IM2.grid(row=0, column=1, padx=10, pady=10, )
IM3 = ttk.Button(IM_FRAME, text="SPI - 3",  command=im3, width=8)
IM3.grid(row=0, column=2, padx=10, pady=10, )
IM4 = ttk.Button(IM_FRAME, text="SPI - 4", command=im4, width=8)
IM4.grid(row=1, column=0, padx=10, pady=10, )
IM5 = ttk.Button(IM_FRAME, text="SPI - 5", command=im5, width=8)
IM5.grid(row=1, column=1, padx=10, pady=10, )
#IM6 = ttk.Button(IM_FRAME,  text="SPI - 6", width=8 )
#IM6.grid(row=1, column=1, padx=10, pady=10, )
#IM7 = ttk.Button(IM_FRAME,  text="SPI - 7", width=8 )
#IM7.grid(row=1, column=2, padx=10, pady=10, )
#IM8 = ttk.Button(IM_FRAME,  text="SPI - 8", width=8 )
#IM8.grid(row=1, column=3, padx=10, pady=10, )

LB_FRAME =ttk.LabelFrame(ram, text="LABORATORIJA", width=400)
LB_FRAME.grid(row= 2, column=1, sticky="news", padx=20, pady=10)

LB1 = ttk.Button(LB_FRAME, text="LB - 1", command=lb1, width=8)
LB1.grid(row=0, column=0, padx=10, pady=10, )
LB2 = ttk.Button(LB_FRAME, text="LB - 2",  command=lb2, width=8)
LB2.grid(row=0, column=1, padx=10, pady=10, )
LB3 = ttk.Button(LB_FRAME, text="LB - 3",  command=lb3, width=8)
LB3.grid(row=0, column=2, padx=10, pady=10, )
LB4 = ttk.Button(LB_FRAME, text="LB - 4", command=lb4, width=8)
LB4.grid(row=1, column=0, padx=10, pady=10, )
LB5 = ttk.Button(LB_FRAME, text="LB - 5", command=lb5, width=8)
LB5.grid(row=1, column=1, padx=10, pady=10, )

zadatak_label = tk.Label(zadatak_info_nalog, text="AKTIVNOST / POSAO", )
zadatak_label.grid(row=0, column=5, padx=10, pady=6, sticky="w")
zadatak_entry = ttk.Entry(zadatak_info_nalog, width=65)
zadatak_entry.grid(row=1, column=5, padx=10, pady=6)

btn_frame =tk.LabelFrame(ram, width=400)
btn_frame.grid(row= 2, column=0, sticky="news", padx=20, pady=10)

unesi_button = ttk.Button(btn_frame, text="UNESI U EVIDENCIJU", width=35, command = Nalog.unesi)
unesi_button.grid(row=0, column=1, columnspan=2, padx=20, pady=10, sticky="ew")
nov_nalog_button = ttk.Button(btn_frame, text="NOV NALOG", width=35, command=reset)
nov_nalog_button.grid(row=0, column=3, columnspan=2, padx=20, pady=10, sticky="ew")

quit_btn = ttk.Button(btn_frame, text="ZATVORI", command=close_app, width=35)
quit_btn.grid(row=0, column=5, padx=20, pady=10, sticky="w")

zapamti_klijenta = ttk.Button(btn_frame, text="UNESI KLIJENTA U BAZU", width=35, command = Nalog.upamti_klijenta)
zapamti_klijenta.grid(row=1, column=1, columnspan=2, padx=20, pady=10, sticky="ew")

print_btn = ttk.Button(btn_frame, text="NALOG U PDF",width=25)
print_btn.grid(row=1, column=3, columnspan=2, padx=20, pady=10, sticky="ew")


for combo in combos:
    combo['values'] = names

window.mainloop()
